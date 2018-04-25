# coding=utf-8
import re,os
import pymysql
import collections
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill,Alignment


def get_data():
    conn = pymysql.connect(host='10.10.9.156', port=3306, user='caohuili', password='chl##098', db='FDA_Carewell',charset='utf8')
    cur = conn.cursor()
    sql_qiang = '''select PatientID,DIID from t_ecg_event WHERE PatientID NOT IN (SELECT PatientID FROM `t_test` WHERE TestType=2 OR TestType=3);'''
    #sql_qiang = '''select PatientID,DIID from t_ecg_event;'''
    sql_qiang_beizhu='select PatientID,CustomDiagnosis from t_patient'

    cur.execute(sql_qiang)
    cur.execute(sql_qiang_beizhu)


    df_qiang = pd.read_sql(sql_qiang, conn).drop_duplicates().dropna()
    df_qiang_beizhu = pd.read_sql(sql_qiang_beizhu, conn).dropna()

    #print(len(set(df_qiang['PatientID'])))

    df_qiang.sort_values('PatientID')

    cur.close()
    conn.commit()
    conn.close()
    return df_qiang, df_qiang_beizhu

def get_excel_result(patient_excel):

    wb=openpyxl.load_workbook('诊断统计.xlsx')
    sheet=wb.get_sheet_by_name('patientid-q')

    df_qiang,df_qiang_beizhu = get_data()

    five_paitient={2869736, 2866471, 2866680, 2863629, 2869735}
    patientids = list(set(df_qiang['PatientID']))
    patientids.sort()
    total_num = len(set(patientids))

    # diids = list(set(df_qiang['DIID']))
    # diids.sort()
    diids=[308, 311, 315, 402, 501, 505, 813, 814, 815, 816, 832, 833, 834, 835, 836, 842, 843, 844, 847, 848, 861, 867, 869, 872, 875]

    diid_num = [0] * len(diids)

    # print(df_carewell[df_carewell['PatientID']==282006])
    line=4
    for patientid in patientids:
        if patientid in five_paitient:
            continue
        #print(patientid)
        q_code_list = list(df_qiang[df_qiang['PatientID']==patientid]['DIID'])
        qiang_beizhu =list(df_qiang_beizhu[df_qiang_beizhu['PatientID'] == patientid]['CustomDiagnosis'])


        for q_code in q_code_list:
            if q_code not in diids:
                continue
            else:
                code_index = diids.index(q_code)
                num = diid_num[code_index]
                sheet.cell(row=4+num,column=code_index+1).value=patientid
                diid_num[code_index]+=1
    col=1
    for diid in diids:
        sheet.cell(row=1, column=col).value = diid
        sheet.cell(row=2, column=col).value = total_num
        sheet.cell(row=3, column=col).value = diid_num[col-1]
        col+=1


    sheet.freeze_panes = 'A2'
    wb.save(patient_excel)

if __name__ == '__main__':
    patient_excel ='20180226_3500-6强大夫诊断结果个数统计.xlsx'
    get_excel_result(patient_excel)
