# coding=utf-8
import pymysql,re
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill


def get_data():
    # conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    conn = pymysql.connect(host='10.10.9.156', port=3306, user='caohuili', password='chl##098', db='StaticECG',
                           charset='utf8')
    cur = conn.cursor()
    sql_1000 = "SELECT PatientID,DoctorAdvice FROM `t_test` where TestType=100 and DoctorAdvice<>''"
    sql_500 = "SELECT PatientID,DoctorAdvice FROM `t_test` where TestType=14 AND PatientID in (SELECT PatientID FROM `t_test` where TestType=100 and DoctorAdvice<>'')"
    sql_q = "SELECT PatientID,DIID FROM `t_ecg_event` where PatientID in (SELECT PatientID FROM `t_test` where TestType=100 and DoctorAdvice<>'')"
    cur.execute(sql_1000)
    cur.execute(sql_500)
    cur.execute(sql_q)
    df_1000 = pd.read_sql(sql_1000, conn)
    df_500 = pd.read_sql(sql_500, conn)
    df_q = pd.read_sql(sql_q, conn)


    print(df_1000.tail(3))

    df_1000 = df_1000[['PatientID', 'DoctorAdvice']]
    df_500 = df_500[['PatientID', 'DoctorAdvice']]
    df_q= df_q[['PatientID', 'DIID']]

    return df_1000,df_500,df_q



def get_ai_rusult():
    df_1000, df_500, df_q = get_data()
    zh_diids_dict = {}
    new_diids_dict = {}


    diids = []

    d = df_1000.to_dict()
    PatientID_dict = d['PatientID']
    DoctorAdvice_dict = d['DoctorAdvice']
    for key,PatientID in PatientID_dict.items():
        doctoradvice = DoctorAdvice_dict[key]
        code = re.findall(r'\d\d\d',doctoradvice)
        diids += code
        for i in code:
            if i in new_diids_dict:
                new_diids_dict[i].append(PatientID)
            else:
                new_diids_dict[i] = [PatientID]

    d_500 = df_500.to_dict()
    zh_PatientID_dict = d_500['PatientID']
    zh_DoctorAdvice_dict = d_500['DoctorAdvice']
    for key,PatientID in zh_PatientID_dict.items():
        doctoradvice = zh_DoctorAdvice_dict[key]
        zh_code = re.findall(r'\d\d\d',doctoradvice)
        diids += zh_code
        for i in zh_code:
            if i in zh_diids_dict:
                zh_diids_dict[i].append(PatientID)
            else:
                zh_diids_dict[i] = [PatientID]

    diids = list(set(diids))
    diids.sort()

    return diids,new_diids_dict,zh_diids_dict

def get_ai_q_rusult():
    df_1000, df_500, df_q = get_data()
    q_diids_dict = {}
    new_diids_dict = {}

    q_more_patient ={2870101, 2870102, 2870103, 2870104, 2870105, 2870106, 2870107, 2870108, 2870109, 2870110, 2870111, 2870112, 2870113, 2870114, 2870115, 2870116, 2870117, 2870118, 2870119, 2870120, 2870121, 2870122, 2870124, 2870125, 2870126, 2870127, 2870128, 2870129, 2870130, 2870131, 2870132, 2870133, 2870134, 2870135, 2870136, 2870137, 2870138, 2870139, 2870140, 2870141, 2870142, 2870143, 2870144, 2870145, 2870146, 2870147, 2870148, 2870149, 2870150, 2870151, 2870152, 2870153, 2870154, 2870155, 2870156, 2870157, 2870158, 2870159, 2870160, 2870161, 2870162, 2870163, 2870164, 2870165, 2870166, 2870167, 2870168, 2870169, 2870170, 2870171, 2870172, 2870173, 2870174, 2870175, 2870176, 2870177, 2870178, 2870179, 2870180, 2870181, 2870182, 2870183, 2870184, 2870185, 2870186, 2870187}

    diids = []

    d = df_1000.to_dict()
    PatientID_dict = d['PatientID']
    DoctorAdvice_dict = d['DoctorAdvice']
    patient_list = []
    for key,PatientID in PatientID_dict.items():
        if int(PatientID) in q_more_patient:
            continue
        patient_list.append(PatientID)
        doctoradvice = DoctorAdvice_dict[key]
        code = re.findall(r'\d\d\d',doctoradvice)
        diids += code
        for i in code:
            if i in new_diids_dict:
                new_diids_dict[i].append(PatientID)
            else:
                new_diids_dict[i] = [PatientID]

    #d_q = df_q.to_dict()
    q_PatientID = list(set(df_q['PatientID']))

    q_PatientID.sort()
    for patientid in q_PatientID:
        q_code_list = list(df_q[df_q['PatientID']==patientid]['DIID'])
        q_code_list= list(map(str,q_code_list))
        diids += q_code_list
        for i in q_code_list:
            if i in q_diids_dict:
                q_diids_dict[i].append(patientid)
            else:
                q_diids_dict[i] = [patientid]

    diids = list(set(diids))
    diids.sort()

    return diids,new_diids_dict,q_diids_dict

def get_excel_result():
    diids, new_diids_dict, zh_diids_dict = get_ai_q_rusult()
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    line = 1
    col = 1
    for diid in diids:
        sheet.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+2)
        sheet.cell(row = line,column=col).value = diid
        try:
            new_diids_list = new_diids_dict[diid]
        except:
            new_diids_list = []
        try:
            zh_diids_list = zh_diids_dict[diid]
        except:
            zh_diids_list = []
        print(new_diids_list)
        print(zh_diids_list)
        match = list(set(new_diids_list)&set(zh_diids_list))

        more = list(set(new_diids_list)-set(zh_diids_list))
        less = list(set(zh_diids_list) - set(new_diids_list))
        match.sort()
        more.sort()
        less.sort()

        sheet.cell(row=line + 1, column=col).value = 'match'
        sheet.cell(row=line + 1, column=col+1).value = 'more'
        sheet.cell(row=line + 1, column=col+2).value = 'less'
        sheet.cell(row=line + 2, column=col).value = len(match)
        sheet.cell(row=line + 2, column=col+1).value = len(more)
        sheet.cell(row=line + 2, column=col+2).value = len(less)

        line_match = 4
        line_more = 4
        line_less = 4

        for i in match:
            sheet.cell(row=line_match, column=col).value = i
            line_match+=1

        for i in more:
            sheet.cell(row=line_more, column=col+1).value = i
            line_more+=1

        for i in less:
            sheet.cell(row=line_less, column=col+2).value = i
            line_less+=1

        col+=3

    sheet.freeze_panes='A4'
    wb.save('20180327静态心电采样率1000-强大夫对比结果.xlsx')

if __name__ == '__main__':
    get_excel_result()
