#coding=utf-8
import pymysql
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill

def get_data():
    #conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    conn = pymysql.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG',charset='utf8')
    cur = conn.cursor()
    sql = 'SELECT * FROM t_report_compare_to_carewell WHERE TestID in (SELECT TestID from t_test WHERE PatientID in (select PatientID from t_patient WHERE  TestType = 17))'
    #sql = 'select * from t_report_compare_to_carewell'
    cur.execute(sql)
    df=pd.read_sql(sql, conn)

    print(df.tail(3))
    df_ai =df[['TestID','AIResult']]
    df_carewell = df[['TestID','CarewellResult']]


    df_carewell.sort_values('TestID')
    print(df_carewell.iloc[13,:])
    #return df_carewell.dropna()
    return df_ai.dropna()

def get_ai_dict():
    #labels=['P波间期', 'PR间期', 'QRS间期', 'QT间期', 'T波间期', 'RV1','SV1','RV5','SV5']
    labels = [ 'PR间期', 'QRS间期', 'QT间期','SV1', 'RV5']
    patientids=['281924', '281925', '281926', '281927', '281928', '281929', '281930', '281931', '281932', '281933', '281934',
     '281935', '281936', '281937', '281938', '281939', '281940', '281941', '281942', '281943', '281946', '281988',
     '281990', '281992']
    testids =['304652', '304653', '304654', '304655', '304656', '304657', '304658', '304659', '304660', '304661', '304662', '304663', '304664', '304665', '304666', '304667', '304668', '304669', '304670', '304671', '304674', '304716', '304718', '304720','304852', '304857', '304859', '304861']
    df_cw = get_data()
    rows = len(df_cw)
    cw_dict = {}
    
    for row in range(rows):
        testid = df_cw.iloc[row,0]
        cwresult = df_cw.iloc[row,1].split(':')
        if testid in cw_dict:
            cw_dict[testid][cwresult[0]]=cwresult[1]
        else:
            cw_dict[testid]={}
            cw_dict[testid][cwresult[0]] = cwresult[1]
    # for i in range(rows):
    #     testid = df_cw.iloc[i, 0]
    #     cwresult = df_cw.iloc[i, 1]
    #     if len(cwresult) < 1:
    #         continue
    #     else:
    #         carewellresult = cwresult.split('：')
    # 
    #     if testid in cw_dict:
    #         cw_dict[testid][carewellresult[0]] = carewellresult[1].replace(' ', '').replace('\n', '')
    #     else:
    #         cw_dict[testid] = {}
    #         cw_dict[testid][carewellresult[0]] = carewellresult[1].replace(' ', '').replace('\n', '')

    wb = openpyxl.load_workbook('E:\\xindian\\静态心电数据库统计\\carewell-人工测试对比结果20180102.xlsx')
    sheet = wb.get_sheet_by_name('result')

    line=5

    for tid in testids:
        col = 2
        for label in labels:
            sheet.cell(row=line, column=col).value = cw_dict[int(tid)][label]
            col+=1
        line+=1
    wb.save('carewell-人工测量数据22.xlsx')

if __name__ == '__main__':
    get_ai_dict()
