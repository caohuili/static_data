import openpyxl
from collections import Counter,defaultdict


def get_lvh_excel():
    #label_list = ['AMI', 'IMI', 'MIX', 'LVH', 'MI+VH', 'Normal', 'OTHERS', 'BVH', 'RVH']
    label_list = ['Normal', 'LVH', 'RVH', 'BVH', 'AMI', 'IMI', 'MIX', 'MI+VH', 'OTHERS']

    wb = openpyxl.load_workbook('赵博士lvh.xlsx')
    sheet = wb.get_sheet_by_name('result')

    df_qiang, df_carewell, codeid_list = get_cse_data()

    label_dict = defaultdict(list)
    for label in label_list:
        label_dict[label]=[0, 0, 0,0,0,0,0,0,0]

    nomarl_num_dict =Counter(list(df_qiang['DIID']))
    cw_num_dict =Counter(list(df_qiang['DIID']))
    patientids = list(df_qiang['PatientID'])
    total_patient = len(patientids)
    for patientid in patientids:
        q_label = list(df_qiang[df_qiang['PatientID']==patientid]['DIID'])[0]
        try:
            c_label = list(df_carewell[df_carewell['PatientID'] == patientid]['DIID'])[0]
        except:
            total_patient -= 1
            continue
        num = label_list.index(c_label)
        label_dict[q_label][num]+=1

    line=2
    for elabel in label_list:
        cse_num_list = label_dict[elabel]
        for i in range(len(cse_num_list)):
            try:
                sheet.cell(row=i+2, column=line).value = round(cse_num_list[i]*100/nomarl_num_dict[elabel],1)
            except:
                sheet.cell(row=line, column=i + 2).value ='-'
        sheet.cell(row=line, column=11).value = round(cw_num_dict[elabel] * 100 / total_patient, 1)
        sheet.cell(row=11, column=line).value = 100
        sheet.cell(row=12, column=line).value = "%d/%d"%(nomarl_num_dict[elabel] ,total_patient)
        sheet.cell(row=12, column=line).alignment = Alignment(horizontal='right',vertical='center')
        line+=1
    sheet.cell(row=11, column=11).value = 100
    sheet.cell(row=12, column=11).value = total_patient

    sheet.freeze_panes = 'A2'
    wb.save('result.xlsx')



if __name__ == '__main__':
    get_lvh_excel()