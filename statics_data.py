import MySQLdb,os,re
import openpyxl


def get_q_patientid():

    conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    cur = conn.cursor()

    sql = "select PatientID from t_patient where DataSource=1"
    cur.execute(sql)
    q_patient_ids = cur.fetchall()
    q_patientid_list=[]
    for patientid in q_patient_ids:
        q_patientid_list.append(patientid[0])
    print(q_patientid_list)

    cur.close()
    conn.commit()
    conn.close()
    return q_patientid_list

def get_q_ecgevent():
    conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    cur = conn.cursor()

    q_patientids = get_q_patientid()
    q_patientid_dict ={}
    for patientid in q_patientids:

        sql = "select DIID from t_ecg_event where PatientID=%s" %(patientid)
        cur.execute(sql)
        events = cur.fetchall()
        events = set(events)

        q_events=[]
        for event in events:
            q_events.append(event[0])
        if 421 in q_events:
            pass
        elif len(q_events)==0:
            pass
        else:
            print(q_events)
            q_patientid_dict[patientid]=q_events
    #print(q_patientid_dict)

    cur.close()
    conn.commit()
    conn.close()
    return q_patientid_dict


def get_ai_ecgevent():
    conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    cur = conn.cursor()

    #q_patientids = get_q_patientid()
    q_patientids = get_q_ecgevent()
    ai_patientid_dict ={}
    for patientid in q_patientids:

        sql = "select DoctorAdvice  from t_test where TestType=5 and PatientID=%s" %(patientid)
        cur.execute(sql)
        events = cur.fetchall()
        events = set(events)
        ai_events=[]
        if bool(events):
            for event in events:
                e=event[0]
                e_list = re.findall('(\d\d\d):',str(e))
                ai_events.append(e)
            print(e_list)
        else:
            e_list = []
        ai_patientid_dict[patientid]=e_list
    #print(ai_patientid_dict)

    cur.close()
    conn.commit()
    conn.close()
    return ai_patientid_dict


def write2excel():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('doctor_q')
    q_dict = get_q_ecgevent()
    line=1
    for pid,q_event in q_dict.items():
        #sheet.cell(row=line,column=1).value='q'
        if bool(q_event):
            for i in range(len(q_event)):
                sheet.cell(row=line, column=1).value = 'q'
                sheet.cell(row=line, column=2).value = str(pid)+'-'+str(q_event[i])
                sheet.cell(row=line, column=3).value = pid
                sheet.cell(row=line, column=4).value = q_event[i]
                line+=1
        else:
            sheet.cell(row=line, column=1).value = 'q'
            sheet.cell(row=line, column=2).value = pid
            line += 1
    wb.save('doctor_q.xlsx')

def write2excel_q():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('doctor_q')
    q_dict = get_q_ecgevent()
    line=1
    for pid,q_event in q_dict.items():
        #sheet.cell(row=line,column=1).value='q'
        if bool(q_event):
            q_event.sort()
            sheet.cell(row=line, column=1).value = 'q'
            sheet.cell(row=line, column=2).value = pid
            sheet.cell(row=line, column=3).value = str(q_event)[1:-1]
            line+=1
        else:
            sheet.cell(row=line, column=1).value = 'q'
            sheet.cell(row=line, column=2).value = pid
            line += 1
    wb.save('doctor_q2.xlsx')

def ai_write2excel():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('doctor_ai')
    ai_dict = get_ai_ecgevent()
    line=1
    for pid,ai_event in ai_dict.items():
        #sheet.cell(row=line,column=1).value='q'
        if bool(ai_event):
            for i in range(len(ai_event)):
                sheet.cell(row=line, column=1).value = 'ai'
                sheet.cell(row=line, column=2).value = str(pid)+'-'+str(ai_event[i])
                sheet.cell(row=line, column=3).value = pid
                sheet.cell(row=line, column=4).value = ai_event[i]
                line+=1
        else:
            sheet.cell(row=line, column=1).value = 'ai'
            sheet.cell(row=line, column=2).value = pid
            line += 1
    wb.save('doctor_ai.xlsx')

def write2excel_ai():
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('doctor_ai')
    ai_dict = get_ai_ecgevent()
    line=1
    for pid,ai_event in ai_dict.items():
        #sheet.cell(row=line,column=1).value='q'
        if bool(ai_event):
            sheet.cell(row=line, column=1).value = 'ai'
            sheet.cell(row=line, column=2).value = pid
            sheet.cell(row=line, column=3).value = ','.join(ai_event)
            line+=1
        else:
            sheet.cell(row=line, column=1).value = 'ai'
            sheet.cell(row=line, column=2).value = pid
            line += 1
    wb.save('doctor_ai2.xlsx')

def match_result():
    qe=['101', '110', '112', '121', '122', '131', '132', '141', '171', '203', '205', '301', '303', '304', '306', '307', '308', '311', '315', '411', '421', '501', '502', '504', '505', '511', '611', '621', '633', '701', '711', '751', '753', '811', '812', '821', '833', '834', '835', '841', '842', '845', '846', '851', '853', '861', '863', '866', '871', '872', '874']

    q_qe_dict = {}
    q_dict = get_q_ecgevent()
    for pid,events in q_dict.items():
        if bool(events):

            for e in events:
                if str(e) in q_qe_dict:
                    q_qe_dict[str(e)].append(pid)
                else:
                    q_qe_dict[str(e)]=[pid]

    ai_qe_dict ={}
    ai_dict = get_ai_ecgevent()
    for pid,events in ai_dict.items():
        if bool(events):

            for e in events:
                if e in ai_qe_dict:
                    ai_qe_dict[e].append(pid)
                else:
                    ai_qe_dict[e]=[pid]

    ai_more=[]
    e_match={}
    e_more={}
    e_less={}
    for ddid,q_patientid in q_qe_dict.items():
        if ddid not in ai_qe_dict:
            ai_more.append([ddid,q_patientid])
        else:
            ai_patientid = ai_qe_dict[ddid]
            e_match[ddid] = list(set(ai_patientid)&set(q_patientid))
            e_more[ddid]=list(set(ai_patientid)-set(q_patientid))
            e_less[ddid]=list(set(q_patientid)-set(ai_patientid))
    #print('ai_more======================================================================')
    #print(ai_more)
    # print('ai_less===================================================================')
    # #ll=list(set(ai_dict)-set(q_dict))
    # dd={}
    # for i in ll:
    #     dd[i] = ai_dict[i]
    #
    #
    # #print(dd)
    return e_match,e_more,e_less

def result_excel():
    #a=['286108', '2861778', '2861813', '2862015', '2862079', '2863353', '2863456', '2864699', '2864700', '2864701', '2866429', '2869161', '2869168', '2869237', '2869267', '2869268', '2869811', '2869805', '281988', '282290', '282314', '282316', '282318', '2861726', '2861922', '2861980', '2862164', '2862238', '2863625', '2863639', '2863661', '2864490', '2864739', '2866193', '2866454', '2870187', '2870186', '2870185', '2870184', '2870183', '2870182', '2870181', '2870180', '2870179', '2870178', '2870177', '2870176', '2870175', '2870174', '2870173', '2870172', '2870171', '2870170', '2870169', '2870168', '2870167', '2870166', '2870165', '2870164', '2870163', '2870162', '2870161', '2870160', '2870159', '2870158', '2870157', '2870156', '2870155', '2870154', '2870153', '2870152', '2870151', '2870150', '2870149', '2870148', '2870147', '2870146', '2870145', '2870144', '2870143', '2870142', '2870141', '2870140', '2870139', '2870138', '2870137', '2870136', '2870135', '2870134', '2870133', '2870132', '2870131', '2870130', '2870129', '2870128', '2870127', '2870126', '2870125', '2870124', '2870122', '2870121', '2870120', '2870119', '2870118', '2870117', '2870116', '2870115', '2870114', '2870113', '2870112', '2870111', '2870110', '2870109', '2870108', '2870107', '2870106', '2870105', '2870104', '2870103', '2870102', '2870101']
    a=[]

    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('result')
    #ai_dict = get_ai_ecgevent()

    col=1
    match, more, less = match_result()
    diids=list(match)
    diids.sort()
    for ddid in diids:
        pmatch_list=match[ddid]
        #line = 1
        #sheet.cell(row=line,column=1).value='q'
        pmore_list = more[ddid]
        pless_list = less[ddid]

        if bool(pmatch_list):
            line=1
            sheet.cell(row=line, column=col).value = ddid
            for p in pmatch_list:
                if str(p) in a:
                    pass
                else:
                    sheet.cell(row=line+1, column=col).value = p
                    line+=1
        else:
            pass

        if bool(pmore_list):
            line = 1
            sheet.cell(row=line, column=col+1).value = ddid
            for p in pmore_list:
                if str(p) not in a:
                    sheet.cell(row=line+1, column=col+1).value = p
                    line+=1
        else:
            pass

        if bool(pless_list):
            line = 1
            sheet.cell(row=line, column=col+2).value = ddid
            for p in pless_list:
                if str(p) not in a:
                    sheet.cell(row=line+1, column=col+2).value = p
                    line+=1
        else:
            pass
        col+=3

    wb.save('张雪门诊诊断统计结果-testtype13.xlsx')


if __name__ == '__main__':
    # get_q_ecgevent()
    # get_ai_ecgevent()
    # write2excel()
    # ai_write2excel()
    #write2excel_q()
    #write2excel_ai()
    result_excel()
    #match_result()