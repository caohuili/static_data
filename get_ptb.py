# coding=utf-8
import re
import pymysql
import collections
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill


def get_data():
    # conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    conn = pymysql.connect(host='10.10.9.156', port=3306, user='caohuili', password='chl##098', db='FDA_PTB',
                           charset='utf8')
    cur = conn.cursor()
    sql_qiang = '''SELECT t_test.PatientID,t_patient.PatientName,t_test.DoctorAdvice FROM `t_test`,t_patient
WHERE t_test.PatientID=t_patient.PatientID and TestType=2'''

    cur.execute(sql_qiang)


    df_qiang = pd.read_sql(sql_qiang, conn)

    cur.close()
    conn.commit()
    conn.close()
    return df_qiang


def get_excel_result():
    #codes = [801,802,803,835,865,836,837,869,803,871,872,110,866,421,875,501,504,505,510,511,512]
    #codes = [801, 802, 803, 835, 865, 836, 837, 869, 871, 872,  875,421, 501, 504, 505, 510, 511, 512]
    code8 = set(range(800, 900))
    code7 = set(range(700,800))
    code5 = set(range(500, 600))
    code3 = set(range(300, 400))
    code1=set(range(100, 300))
    codes = code1|code3|code5|code7|code8
    wb = openpyxl.load_workbook('诊断统计.xlsx')
    sheet = wb.get_sheet_by_name('patientid')
    fill1 = PatternFill("solid", fgColor="F4AF85")
    fill2 = PatternFill("solid", fgColor="FED966")
    fill_color=[fill1,fill2]
    df_qiang = get_data()

    patientids = list(set(df_qiang['PatientID']))
    patientids.sort()

    line = 2
    patient_dict ={}
    sid_dict ={}
    for patientid in patientids:

        advice_text = list(df_qiang[df_qiang['PatientID']==patientid]['DoctorAdvice'])[0]
        patient =list(df_qiang[df_qiang['PatientID']==patientid]['PatientName'])[0]
        print(patient)

        advice = '|'+str(advice_text)
        advice=advice.replace('||', '|').replace(': ', ':').replace('    ','')
        advice_list = advice.split('|')
        print(advice)
        advice_list=re.findall(r'\|(\d\d\d):',advice)
        advice_set=set(map(int,advice_list))

        pid = patient[:10]
        sid = patient[11:19]
        rid = patient[17:-4]
        if sid in sid_dict:
            sid_dict[sid].append(patientid)
        else:
            sid_dict[sid]=[patientid]

        if pid in patient_dict:
            if sid in patient_dict[pid]:

                patient_dict[pid][sid]=patient_dict[pid][sid] | advice_set
            else:

                patient_dict[pid][sid]=advice_set
        else:
            patient_dict[pid]={}
            patient_dict[pid][sid]=advice_set
    line=2
    patient_list = list(patient_dict.keys())
    patient_list.sort()
    for k in range(len(patient_list)):
        v=patient_dict[patient_list[k]]

        for sid,code_set in v.items():

            code_list= list(code_set&codes)
            #code_list=list(code_set)
            code_list.sort()

            c_text = ''
            #for c in code_list:
                #c_text += (str(c)+',')
                # sheet.cell(row=line, column=1).value = patient_list[k]
                # sheet.cell(row=line, column=2).value = sid
                # sheet.cell(row=line,column=3).value=c
                # sheet.cell(row=line, column=1).fill = fill_color[k%2]
                # sheet.cell(row=line, column=2).fill = fill_color[k%2]
                # sheet.cell(row=line, column=3).fill = fill_color[k%2]
            sheet.cell(row=line, column=1).value = patient_list[k]
            sheet.cell(row=line, column=2).value = sid
            sheet.cell(row=line,column=3).value=str(code_list)[1:-1]
            sheet.cell(row=line, column=4).value = max(sid_dict[sid])
            sheet.cell(row=line, column=5).value = min(sid_dict[sid])
            line+=1

    sheet.freeze_panes = 'A2'
    wb.save('20180202_ptb_结果统计.xlsx')

def match():

    anterior=[731,741,751,761,771,781,791]
    infero_postero_lateral =[735,745,755,733,743,753,763,773,732,742,752,762,772]#
    antero_septal=[734,744,754,764]
    inferior=[733,743,753,763,773]
    infero_lateral=[737,747,757]
    antero_lateral=[736,746,756]
    lateral=[732,742,752,762,772]
    postero_lateral=[735,745,755,732,742,752,762,772]#
    posterior=[735,745,755]
    infero_posterior=[735,745,755,733,743,753,763,773]#
    antero_septo_lateral=[734,744,754,764,732,742,752,762,772]#

    acute = [infero_lateral,anterior,infero_postero_lateral,antero_septal,inferior,antero_lateral,lateral,postero_lateral,posterior,infero_posterior,antero_septo_lateral]
    acute_label=['infero-lateral', 'anterior', 'infero-postero-lateral', 'antero-septal', 'inferior', 'antero-lateral', 'lateral', 'postero-lateral', 'posterior', 'infero-posterior', 'antero-septo-lateral']
    wb=openpyxl.load_workbook('20180202_ptb_结果统计.xlsx')
    sheet=wb.get_sheet_by_name('patientid')
    lines = sheet.max_row
    print(lines)
    for i in range(2,lines):
        k = sheet.cell(row=i,column=6).value
        try:
            k=k.replace('\n\n','').replace('\n','').replace('_x000D_','')
            print(k)
        except:
            print(k)

        s = sheet.cell(row=i,column=9).value
        if bool(s):
            c_list = s.replace(' ','').split(',')
            c_list = list(map(int, c_list))
        else:
            c_list=[]

        if k =='infero-postero-lateral':
            inf = len(set(inferior) & set(c_list))
            pos = len(set(posterior) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            sheet.cell(row=i, column=8).value = str(infero_postero_lateral)[1:-1]
            if inf != 0 and pos != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k =='postero-lateral':
            sheet.cell(row=i, column=8).value = str(postero_lateral)[1:-1]
            pos = len(set(posterior) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            if pos != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k =='infero-posterior':
            sheet.cell(row=i, column=8).value = str(infero_posterior)[1:-1]
            pos = len(set(posterior) & set(c_list))
            inf = len(set(inferior) & set(c_list))
            if pos != 0 and inf != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k == 'antero-septo-lateral':
            sheet.cell(row=i, column=8).value = str(antero_septo_lateral)[1:-1]
            ant = len(set(antero_septal) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            if ant != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'

        elif k not in acute_label and len(c_list)==0:
            sheet.cell(row=i, column=7).value = '匹配'
        elif k not in acute_label:
            continue
        else:
            a_list = acute[acute_label.index(k)]

            r_set = set(c_list)&set(a_list)
            sheet.cell(row=i, column=8).value=str(a_list)[1:-1]
            if len(r_set)==0:
                sheet.cell(row=i, column=7).value ='不匹配'
            else:
                sheet.cell(row=i, column=7).value = '匹配'


    wb.save('result120180222-3.xlsx')

if __name__ == '__main__':
    # ptb:
    # 1、get_ptb.py得到数据库中数据与physiobank_database_result的DIID码对比；
    # 2、手动统计"Myocardial infarction-1"，"Myocardial infarction-2"，"Myocardial infarction-3"，
    # "Healthy control"，"Dysrhythmia"，"Hypertrophy"，"Bundle branch block"事件的TP FP FN的数值；
    # 3、手动写入sen - pos的sheet页中，自动计算出SENS( %) SPEC( %) PPV( %)    NPV( %)    PREV
    # PREV( %)各个比值。

    match()

    #get_excel_result()
    # get_excel_patient_result(codeid_list)
    # get_excel_senpos_result(codeid_list)