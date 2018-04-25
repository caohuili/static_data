# coding=utf-8
import re,os
import pymysql
import collections
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill,Alignment


def get_data():
    conn = pymysql.connect(host='10.10.9.156', port=3306, user='caohuili', password='chl##098', db='FDA_Carewell',
                           charset='utf8')
    cur = conn.cursor()
    sql_qiang = '''select PatientID,DIID from t_ecg_event WHERE PatientID IN (SELECT PatientID FROM `t_test` WHERE TestType=2 or TestType=3);'''
    sql_carewell = 'select PatientID,DoctorAdvice from t_test WHERE TestType=2 or TestType=3'
    sql_qiang_beizhu='select PatientID,CustomDiagnosis from t_patient'


    cur.execute(sql_qiang)
    cur.execute(sql_carewell)
    cur.execute(sql_qiang_beizhu)


    df_qiang = pd.read_sql(sql_qiang, conn).drop_duplicates()
    df_carewell = pd.read_sql(sql_carewell, conn).drop_duplicates()
    df_qiang_beizhu = pd.read_sql(sql_qiang_beizhu, conn).dropna()

    print(list(df_qiang_beizhu[df_qiang_beizhu['PatientID'] == 282006]['CustomDiagnosis']))
    df_qiang.sort_values('PatientID')
    df_carewell.sort_values('PatientID')
    print(len(df_carewell),len(df_qiang))
    #df_qiang.set_index('PatientID')
    print(list(df_qiang[df_qiang['PatientID']==282006]['DIID']))
    cur.close()
    conn.commit()
    conn.close()
    return df_qiang, df_carewell,df_qiang_beizhu


def get_cse_data():
    pass

def get_excel_patient_result(codeid_list,patient_excel):

    wb=openpyxl.load_workbook('诊断统计.xlsx')
    sheet=wb.get_sheet_by_name('patientid')

    df_qiang, df_carewell = get_cse_data()

    codeid_dict = collections.defaultdict(list)
    for codeid in codeid_list:
        codeid_dict[codeid]=[0, 0, 0]

    patientids = list(set(df_qiang['PatientID']))
    patientids.sort()
    total_patient = len(patientids)
    #print(patientids[-10:-1])

    line=4
    for patientid in patientids:
        q_code_list = list(df_qiang[df_qiang['PatientID']==patientid]['DIID'])
        try:
            c_code_list = list(df_carewell[df_carewell['PatientID']==patientid]['DIID'])
        except:
            total_patient -=1
            continue

        c_code_set = set(c_code_list)
        match = set(q_code_list)&c_code_set
        q_less = c_code_set-set(q_code_list)
        q_more = set(q_code_list)-c_code_set

        col=1
        for codeid in codeid_list:
            if codeid in match:
                codeid_dict[codeid][0]+=1
                sheet.cell(row=codeid_dict[codeid][0]+3, column=col).value = patientid
            if codeid in q_more:
                codeid_dict[codeid][1] += 1
                sheet.cell(row=codeid_dict[codeid][1]+3, column=col+1).value = patientid
            if codeid in q_less:
                codeid_dict[codeid][2] += 1
                sheet.cell(row=codeid_dict[codeid][2]+3, column=col+2).value = patientid
            col+=3

    col_num = 1

    label = ['TP','FN','FP']
    for codeid in codeid_list:
        code_num_list = codeid_dict[codeid]
        sheet.merge_cells(start_row = 1,end_row=1,start_column=col_num,end_column=col_num+2)
        sheet.cell(row=1,column=col_num).value=codeid
        for i in range(3):
            sheet.cell(row=2, column=col_num+i).value = label[i]
            sheet.cell(row=3, column=col_num+i).value = code_num_list[i]
        col_num+=3
    sheet.freeze_panes='A2'
    wb.save(patient_excel)
    return total_patient,codeid_dict

def get_excel_senpos_result(codeid_list,patient_excel,sen_pos_excel):

    total_patient, codeid_dict = get_excel_patient_result(codeid_list,patient_excel)

    wb=openpyxl.load_workbook(patient_excel)
    sheet=wb.get_sheet_by_name('sen-pos')

    line=2
    for codeid in codeid_list:
        sheet.cell(row=line, column=1).value = codeid
        code_num_list = codeid_dict[codeid]

        tp=code_num_list[0]
        fn = code_num_list[1]
        fp = code_num_list[2]
        tn = total_patient-tp-fn-fp

        all_num_list = [tp, tn, fp, fn, total_patient]
        col = 2

        for num in all_num_list:
            sheet.cell(row=line, column=col).value = num
            col+=1
            if col==7:

                sheet.cell(row=line, column=7).value = zero_division('=ROUND(B%d*100/(B%d+E%d),2)' % (line, line, line), tp, (tp + fn))
                sheet.cell(row=line, column=8).value = zero_division('=ROUND(C%d*100/(C%d+D%d),2)' % (line, line, line), tn, (tn + fp))
                sheet.cell(row=line, column=9).value = zero_division('=ROUND(B%d*100/(B%d+D%d),2)' % (line, line, line), tp, (tp + fp))
                sheet.cell(row=line, column=10).value = zero_division('=ROUND(C%d*100/(C%d+E%d),2)' % (line, line, line), tn, (tn + fn))
                sheet.cell(row=line, column=11).value = '=SUM(B%d,E%d)&"/"&F%d' % (line, line, line)
                sheet.cell(row=line, column=11).alignment = Alignment(horizontal='right', vertical='center')
                sheet.cell(row=line, column=12).value = zero_division('=ROUND(SUM(B%d,E%d)*100/F%d,2)' % (line, line, line), tp + fn, total_patient)
        line+=1

    sheet.freeze_panes = 'A2'
    wb.save(sen_pos_excel)
    if os.path.exists(patient_excel):
        os.remove(patient_excel)


def zero_division(s,x,y):
    try:
        z=x/y
    except:
        s='-'
    return s


if __name__ == '__main__':
    patient_excel ='20180127_强大夫_carewell_诊断结果个数统计.xlsx'
    sen_pos_excel ='20180127_强大夫_carewell_诊断结果敏感阳性率统计.xlsx'
    #get_data()
    codeid_list=[201,202,203,204,205,206]
    get_excel_senpos_result(codeid_list,patient_excel,sen_pos_excel)
