# coding=utf-8
import re
import pymysql
import collections
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill


def match():

    code_hc = set(range(100,300))
    code_bb = set(range(500, 600))
    code_d = set(range(800, 900))
    code_h = set(range(300, 400))


    anterior=[731,741,751,761,771,781,791]
    infero_postero_lateral =[735,745,755,733,743,753,763,773,732,742,752,762,772]#
    antero_septal=[734,744,754,764]
    inferior=[733,743,753,763,773]
    infero_lateral=[737,747,757]
    antero_lateral=[736,746,756]
    lateral=[732,742,752,762,772]
    postero_lateral=[735,745,755,732,742,752,762,772]#
    posterior=[735,745,755]
    infero_posterior=[735,745,755,733,743,753,763,773]#
    antero_septo_lateral=[734,744,754,764,732,742,752,762,772]#

    wb=openpyxl.load_workbook('result120180202-2.xlsx')
    sheet=wb.get_sheet_by_name('patientid')
    lines = sheet.max_row
    print(lines)
    k3 = ['infero-lateral', 'infero-postero-lateral', 'postero-lateral']
    k1 = [ 'anterior', 'antero-septal','antero-lateral', 'lateral', 'antero-septo-lateral']
    k2 = [ 'inferior','posterior', 'infero-posterior']
    k4 = ['no']

    acute_label=k1+k2+k3+k4
    for i in range(2,lines+1):
        e = sheet.cell(row=i, column=5).value
        k = sheet.cell(row=i,column=6).value
        try:
            e=e.replace('\n\n','').replace('\n','')
            k = k.replace('\n\n', '').replace('\n', '').replace('_x000D_', '')
            print(e,k)
        except:
            print(k)

        ptb = sheet.cell(row=i,column=9).value
        if bool(ptb):
            c_list = ptb.replace(' ','').split(',')
            c_list = list(map(int, c_list))
        else:
            c_list=[]

        if e=='Myocardial infarction':
            if k in k1:
                e_diid = sheet.cell(row=i,column=8).value
                e_list = e_diid.replace(' ', '').split(',')
                e_list = list(map(int, e_list))



        if k =='infero-postero-lateral':
            inf = len(set(inferior) & set(c_list))
            pos = len(set(posterior) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            sheet.cell(row=i, column=8).value = str(infero_postero_lateral)[1:-1]
            if inf != 0 and pos != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k =='postero-lateral':
            sheet.cell(row=i, column=8).value = str(postero_lateral)[1:-1]
            pos = len(set(posterior) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            if pos != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k =='infero-posterior':
            sheet.cell(row=i, column=8).value = str(infero_posterior)[1:-1]
            pos = len(set(posterior) & set(c_list))
            inf = len(set(inferior) & set(c_list))
            if pos != 0 and inf != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'
        elif k == 'antero-septo-lateral':
            sheet.cell(row=i, column=8).value = str(antero_septo_lateral)[1:-1]
            ant = len(set(antero_septal) & set(c_list))
            lat = len(set(lateral) & set(c_list))
            if ant != 0 and lat != 0:
                sheet.cell(row=i, column=7).value = '匹配'
            else:
                sheet.cell(row=i, column=7).value = '不匹配'

        elif k not in acute_label and len(c_list)==0:
            sheet.cell(row=i, column=7).value = '匹配'
        elif k not in acute_label:
            continue
        else:
            a_list = acute[acute_label.index(k)]

            r_set = set(c_list)&set(a_list)
            sheet.cell(row=i, column=8).value=str(a_list)[1:-1]
            if len(r_set)==0:
                sheet.cell(row=i, column=7).value ='不匹配'
            else:
                sheet.cell(row=i, column=7).value = '匹配'


    wb.save('result120180222-2.xlsx')

if __name__ == '__main__':
    # get_data()
    match()
    #get_excel_result()
    # get_excel_patient_result(codeid_list)
    # get_excel_senpos_result(codeid_list)