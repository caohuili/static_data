import pymysql
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill

def get_data():
    #conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    conn = pymysql.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG',charset='utf8')
    cur = conn.cursor()
    sql = 'SELECT * FROM t_report_compare_to_carewell'
    cur.execute(sql)
    df=pd.read_sql(sql, conn)

    print(df.tail(3))
    df_ai =df[['TestID','AIResult']]
    df_carewell = df[['TestID','CarewellResult']]

    return df_ai.dropna(),df_carewell.dropna()

def get_ai_dict():
    labels=['平均心率', 'PR间期', 'RR间期', 'QRS间期', 'QT间期', 'QTc间期', 'QRS ax', 'RV5', 'SV1']
    df_ai, df_carewell = get_data()
    ai_dict ={}
    ai_rows = len(df_ai)
    for row in range(ai_rows):
        testid = df_ai.iloc[row,0]
        airesult = df_ai.iloc[row,1].split(':')
        if '算法分析' in airesult[0]:
            continue

        if testid in ai_dict:
            ai_dict[testid][airesult[0]]=airesult[1]
        else:
            ai_dict[testid]={}
            ai_dict[testid][airesult[0]] = airesult[1]



    carewell_dict = {}
    carewell_rows = len(df_carewell)
    #print(df_carewell.tail(4))
    for row in range(carewell_rows):
        testid = df_carewell.iloc[row, 0]
        cwresult = df_carewell.iloc[row, 1]
        if len(cwresult)<1:
            continue
        else:
            carewellresult=cwresult.split('：')

        if testid in carewell_dict:
            carewell_dict[testid][carewellresult[0]] = carewellresult[1].replace(' ','').replace('\n','')
        else:
            carewell_dict[testid] = {}
            carewell_dict[testid][carewellresult[0]] = carewellresult[1].replace(' ','').replace('\n','')

    fill = PatternFill("solid", fgColor="F4AF85")
    #fill_tl = PatternFill("solid", fgColor="FED966")


    wb = openpyxl.load_workbook('ai_carewell.xlsx')
    sheet=wb.get_sheet_by_name('Sheet2')
    line=2
    for k,v in ai_dict.items():
        carewell_v = carewell_dict[k]

        for label in labels:
            sheet.cell(row=line, column=1).value = k
            sheet.cell(row=line, column=2).value = label
            sheet.cell(row=line,column=3).value = v[label]
            sheet.cell(row=line, column=4).value = carewell_v[label]
            if v[label]==carewell_v[label]:
                sheet.cell(row=line, column=3).fill = fill
                sheet.cell(row=line, column=4).fill = fill
            line+=1

    wb.save('ai_carewell_result5.xlsx')


    # wb = openpyxl.load_workbook('ai_carewell.xlsx')
    # sheet=wb.get_sheet_by_name('Sheet1')
    # line=2
    # for k,v in ai_dict.items():
    #     carewell_v = carewell_dict[k]
    #     sheet.cell(row=line, column=2).value = k
    #     sheet.cell(row=line, column=1).value='ai'
    #     sheet.cell(row=line + 1, column=2).value = k
    #     sheet.cell(row=line+1, column=1).value ='carewell'
    #     col=3
    #     for label in labels:
    #         sheet.cell(row=line,column=col).value = v[label]
    #         sheet.cell(row=line+1, column=col).value = carewell_v[label]
    #         col+=1
    #     line+=2
    # wb.save('ai_carewell_result3.xlsx')





if __name__ == '__main__':
    get_ai_dict()