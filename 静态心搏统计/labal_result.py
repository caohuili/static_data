import openpyxl,pprint
import xlrd


def get_label(q_sheetname,ai_sheetname):
    wbrd = xlrd.open_workbook('E:\\xindian\\静态心电数据库统计\\静态心搏统计\\静态心搏统计.xlsx')
    wb =  openpyxl.load_workbook('E:\\xindian\\静态心电数据库统计\\静态心搏统计\\静态心搏统计.xlsx')
    #wb =  openpyxl.load_workbook('E:\\xindian\\静态心电数据库统计\\静态心搏统计\\AI心搏比较.xlsx')
    # sheet_ai = wb.get_sheet_by_name('AItl-1')
    # sheet_q = wb.get_sheet_by_name('q')

    sheet_aird = wbrd.sheet_by_name(ai_sheetname)
    sheet_qrd = wbrd.sheet_by_name(q_sheetname)

    sheet_ai = wb.get_sheet_by_name(ai_sheetname)
    sheet_q = wb.get_sheet_by_name(q_sheetname)
    ai_lines = sheet_ai.max_row
    q_lines = sheet_q.max_row
    ai={}
    q={}

    for i in range(2,ai_lines+1):
        patientid = sheet_ai.cell(row = i,column=1).value
        rpos = int(sheet_ai.cell(row = i,column=2).value)
        # rpos=int(rpos)
        rlabel = sheet_ai.cell(row = i,column=3).value
        if patientid in ai:

            pass
        else:
            ai[patientid]={}
        ai[patientid][rpos]=rlabel

    q_more_label_dict = {'N_RB': "N", 'N_LAE': "N", 'N_RAE': "N", 'Af_LVH': "Af", 'Af_CLB': "Af", 'Af_RB': "Af"}
    for i in range(2,q_lines+1):
        patientid = sheet_q.cell(row = i,column=1).value
        rpos = int(sheet_q.cell(row = i,column=2).value)
        # rpos = int(rpos)
        rlabel = sheet_q.cell(row = i,column=3).value
        if rlabel in q_more_label_dict:
            rlabel = q_more_label_dict[rlabel]
        if patientid in q:
            pass
        else:
            q[patientid]={}
        q[patientid][rpos]=rlabel

    q_labels = sheet_qrd.col_values(2)
    ai_labels = sheet_aird.col_values(2)
    return ai,q,q_labels,ai_labels


    # with open('ai_r_dict.py', 'w', encoding='utf-8') as f:
    #     f.write(('ecg_r_dict= ') + pprint.pformat(ai))
    # with open('q_r_dict.py', 'w', encoding='utf-8') as fq:
    #     fq.write(('ecg_r_dict= ') + pprint.pformat(q))


def label_resutl(q_sheetname,ai_sheetname):
    ai_dict,q_dict,q_labels,ai_labels = get_label(q_sheetname,ai_sheetname)
    match={}
    more={}
    less={}
    error={}
    # all_label=['N', 'Af','AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'Af_LVH', 'N_CLB', 'Af_CRB', 'N_RB', 'V', 'Af_CLB', 'SE',
    #  'N_LAE', 'N_PS', 'JE', 'A', 'VE', 'AT']
    all_label = ['N', 'Af','AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'N_CLB', 'Af_CRB', 'V', 'SE',
                'N_PS', 'JE', 'A', 'VE', 'AT','VT']

    q_more_label=['N_RB','N_LAE','N_RAE','Af_LVH','Af_CLB','Af_RB']
    #q_more_label_dict = {'N_RB':"N", 'N_LAE':"N", 'N_RAE':"N", 'Af_LVH':"Af", 'Af_CLB':"Af", 'Af_RB':"Af"}
    q_total={}
    ai_total={}
    match_label = {}

    #new_q_labels = [q_more_label_dict[x] if x in q_more_label_dict else x for x in q_labels]

    for lab in all_label:
        q_total[lab]=q_labels.count(lab)
        ai_total[lab] = ai_labels.count(lab)
        match_label[lab]=0

    for i in q_more_label:
        if 'N_' in i:
            q_total['N']+=q_labels.count(i)

        if 'Af_' in i:
            q_total['Af']+=q_labels.count(i)

    for patientid,label_dict in q_dict.items():
        match[patientid] = []
        more[patientid]=[]
        less[patientid]=[]
        error[patientid]=[]
        ai_label_dict = ai_dict.get(patientid)
        if ai_label_dict is None:
            continue

        ql = []


        ####################
        q_rpos_list = list(label_dict)
        ai_rpos_list = list(ai_label_dict)
        q_rpos_list.sort()
        ai_rpos_list.sort()

        if (max(q_rpos_list) - max(ai_rpos_list)) > 150:
            q_total[label_dict[max(q_rpos_list)]] -= 1
            label_dict.pop(max(q_rpos_list))
            q_rpos_list.pop()

        for qpos in q_rpos_list[:2]:

            # if label_dict[qpos] not in all_label:
            #     label_dict.pop(qpos)
            if (min(ai_rpos_list)-qpos)>100:
                if label_dict[qpos] in all_label:
                    q_total[label_dict[qpos]] -= 1
                label_dict.pop(qpos)
            else:
                pass


        #####################
        for aipos,ailabel in ai_label_dict.items():
            m=[]
            #total[ailabel]+=1

            for pos, label in label_dict.items():
                if abs(aipos-pos)<11:
                    if label==ailabel:
                        match[patientid].append([pos,ailabel])
                        ql.append(pos)
                        m.append(aipos)
                        match_label[ailabel]+=1

                    else:
                        error[patientid].append([aipos,ailabel,pos,label])
                        m.append(aipos)
                        ql.append(pos)
                else:
                    pass
            if len(m)==0:
                more[patientid].append([aipos,ailabel])
                #print(aipos,ailabel)
            else:
                pass
        lp=set(label_dict)-set(ql)
        if len(lp)!=0 :
            for l in lp:
                less[patientid].append([l,label_dict[l]])
                #print(l,label_dict[l])


        #print(more,less)
    return match,more,less,error,q_total,ai_total,match_label

def excel_result(q_sheetname):
    wb = openpyxl.load_workbook('统计模板.xlsx')
    sheet_num = wb.get_sheet_by_name('个数-result')

    ai_sheetnames = ['6','7','9','10','8','11']
    for r_col in range(3,9):
        #excel_result(q_sheetname,ai_sheetnames[r_col-3],r_col)
        print(ai_sheetnames[r_col-3]+"完成")

        sheet = wb.get_sheet_by_name(ai_sheetnames[r_col-3])
        match, more, less, error,q_total,ai_total,match_label = label_resutl(q_sheetname,ai_sheetnames[r_col-3])

        matchline=4
        moreline = 4
        lessline = 4
        errorline = 4

        # matchnum=0
        # morenum = 0
        # lessnum = 0
        # errornum = 0
        # all_label=['N', 'Af','AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'Af_LVH', 'N_CLB', 'Af_CRB', 'N_RB', 'V', 'Af_CLB', 'SE',
        #  'N_LAE', 'N_PS', 'JE', 'A', 'VE', 'AT']
        all_label = ['N', 'Af','AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'N_CLB', 'Af_CRB', 'V', 'SE',
                    'N_PS', 'JE', 'A', 'VE', 'AT','VT']
        for pid,matchlist in match.items():
            #sheet.cell(row=2,column=1).value=len(matchlist)
            for matchpr in matchlist:
                sheet.cell(row=matchline, column=1).value = pid
                sheet.cell(row=matchline, column=2).value =matchpr[0]
                sheet.cell(row=matchline, column=3).value = matchpr[1]
                matchline+=1

        for pid,errorlist in error.items():
            #sheet.cell(row=2,column=1).value=len(matchlist)
            for errorpr in errorlist:
                sheet.cell(row=errorline, column=4).value = pid
                sheet.cell(row=errorline, column=5).value =errorpr[2]
                sheet.cell(row=errorline, column=6).value = errorpr[3]
                sheet.cell(row=errorline, column=7).value =errorpr[0]
                sheet.cell(row=errorline, column=8).value = errorpr[1]
                errorline+=1

        for pid,morelist in more.items():
            #sheet.cell(row=2,column=1).value=len(morelist)
            for morepr in morelist:
                sheet.cell(row=moreline, column=9).value = pid
                sheet.cell(row=moreline, column=10).value =morepr[0]
                sheet.cell(row=moreline, column=11).value = morepr[1]
                moreline+=1

        for pid,lesslist in less.items():
            #sheet.cell(row=2,column=1).value=len(lesslist)
            for lesspr in lesslist:
                sheet.cell(row=lessline, column=12).value = pid
                sheet.cell(row=lessline, column=13).value =lesspr[0]
                sheet.cell(row=lessline, column=14).value = lesspr[1]
                lessline+=1
        sheet.cell(row=2, column=1).value = matchline-4
        sheet.cell(row=2, column=4).value = errorline - 4
        sheet.cell(row=2, column=9).value = moreline - 4
        sheet.cell(row=2, column=12).value = lessline - 4



        labnum_line = 3
        for lab in all_label:
            sheet_num.cell(row=labnum_line, column=2).value = q_total[lab]
            sheet_num.cell(row=labnum_line + 19, column=2).value = q_total[lab]
            sheet_num.cell(row=labnum_line, column=r_col).value = ai_total[lab]
            sheet_num.cell(row=labnum_line+19, column=r_col).value = match_label[lab]
            # if total[lab] ==0:
            #     pass
            # else:
            #     sheet_num.cell(row=labnum_line+17, column=3).value = match_label[lab]/total[lab]
            labnum_line+=1
    wb.save('20180418_采样率张玥强大夫250对比结果dd.xlsx')

if __name__ == '__main__':
    #get_label()
    #label_resutl()
    q_sheetname = 'q'

    excel_result(q_sheetname)
