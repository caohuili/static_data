import openpyxl

def compare_qrs_rpos():
    wb = openpyxl.load_workbook('AI心搏比较.xlsx')
    sheet5 = wb.get_sheet_by_name('5-LABEL')
    sheet13 = wb.get_sheet_by_name('13-LABEL')
    lines_five = sheet5.max_row
    lines_tirtheen = sheet13.max_row
    dict5 = {}
    dict13={}
    for line in range(2,lines_five+1):
        pid = sheet5.cell(row=line,column=1).value
        rpos = sheet5.cell(row=line,column=2).value
        if pid in dict5:
            dict5[pid].append(rpos)
        else:
            dict5[pid]=[rpos]

    for line in range(2,lines_tirtheen+1):
        pid = sheet13.cell(row=line,column=2).value
        rpos = sheet13.cell(row=line,column=3).value
        if pid in dict13:
            dict13[pid].append(rpos)
        else:
            dict13[pid]=[rpos]

    match={}
    match5={}
    move={}
    move60={}
    move80 = {}
    move5 ={}
    more={}
    less={}
    more60={}
    less60={}
    for key,list5 in dict5.items():
        list13 = dict13[key]
        # match[key]=list(set(list5)&set(list13))
        # notmatch5 = list(set(list5)-set(list13))
        # notmatch13 = list(set(list13) - set(list5))
        # notmatch = notmatch5+notmatch13
        # notmatch.sort()
        # move[key]=[]
        # move60[key] = []
        # move80[key] = []
        # more60[key]=[]
        # less60[key]=[]
        # if len(notmatch5)!=0 and len(notmatch13)!=0:
        #     for i in range(1,len(notmatch)):
        #         if (notmatch[i]-notmatch[i-1])<41:
        #             if notmatch[i] in notmatch13:
        #                 move[key].append(notmatch[i])
        #             else:
        #                 move[key].append(notmatch[i-1])
        #
        #         elif 40<(notmatch[i] - notmatch[i - 1]) < 61:
        #             if notmatch[i] in notmatch13:
        #                 move60[key].append(notmatch[i])
        #             else:
        #                 move60[key].append(notmatch[i - 1])
        #
        #         # elif (notmatch[i] - notmatch[i - 1])>61:
        #         #     if notmatch[i] in notmatch13:
        #         #         more60[key].append(notmatch[i])
        #         #     else:
        #         #         less60[key].append(notmatch[i - 1])
        #
        # elif len(notmatch5)!=0 and len(notmatch13)==0:
        #     less[key]=notmatch5
        # elif len(notmatch5)==0 and len(notmatch13)!=0:
        #     more[key]=notmatch13
        match[key]=[]
        match5[key]=[]

        for i in range(len(list5)):
            for j in range(len(list13)):
                if abs(list13[j]-list5[i])<41:
                    match[key].append(list13[j])
                    match5[key].append(list5[i])

        more[key]=list(set(list13)-set(match[key]))
        less[key]=list(set(list5)-set(match5[key]))

    sheet = wb.create_sheet('result')

    match_line=2
    for key,pos_list in match.items():
        for r in pos_list:
            sheet.cell(row=match_line,column=1).value=key
            sheet.cell(row=match_line, column=2).value = r
            match_line+=1
            
    # move_line=2
    # for key,pos_list in move.items():
    #     for r in pos_list:
    #         sheet.cell(row=move_line,column=3).value=key
    #         sheet.cell(row=move_line, column=4).value = r
    #         move_line+=1
    #
    # move_line60 = 2
    # for key, pos_list in move60.items():
    #     for r in pos_list:
    #         sheet.cell(row=move_line60, column=9).value = key
    #         sheet.cell(row=move_line60, column=10).value = r
    #         move_line60 += 1
            

    # move_line80 = 2
    # for key, pos_list in move80.items():
    #     for r in pos_list:
    #         sheet.cell(row=move_line80, column=11).value = key
    #         sheet.cell(row=move_line80, column=12).value = r
    #         move_line80 += 1
            
    more_line=2
    for key,pos_list in more.items():
        for r in pos_list:
            sheet.cell(row=more_line,column=5).value=key
            sheet.cell(row=more_line, column=6).value = r
            more_line+=1
            
    less_line=2
    for key,pos_list in less.items():
        for r in pos_list:
            sheet.cell(row=less_line,column=7).value=key
            sheet.cell(row=less_line, column=8).value = r
            less_line+=1

    wb.save('QRS2.xlsx')

if __name__ == '__main__':
    compare_qrs_rpos()