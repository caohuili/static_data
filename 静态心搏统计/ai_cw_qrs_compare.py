import openpyxl
from openpyxl.styles import PatternFill


def get_label():
    wb = openpyxl.load_workbook('E:\\xindian\\静态心电数据库统计\\静态心搏统计\\静态心搏统计.xlsx')
    sheet_ai = wb.get_sheet_by_name('AI')
    sheet_cw = wb.get_sheet_by_name('CW')
    ai_lines = sheet_ai.max_row
    cw_lines = sheet_cw.max_row
    ai = {}
    cw = {}


    for i in range(2, ai_lines + 1):
        patientid = sheet_ai.cell(row=i, column=1).value
        rpos = sheet_ai.cell(row=i, column=2).value

        if patientid in ai:
            pass
        else:
            ai[patientid] = []
        ai[patientid].append(rpos)

    for i in range(2, cw_lines + 1):
        patientid = sheet_cw.cell(row=i, column=1).value
        rpos = sheet_cw.cell(row=i, column=3).value
        if patientid in cw:
            pass
        else:
            cw[patientid] = []
        cw[patientid].append(rpos)


    return ai, cw




def get_result(date):
    wb = openpyxl.load_workbook('统计模板2.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    sheet2 = wb.get_sheet_by_name('Sheet2')

    ai_dict,cw_dict = get_label()
    fill_ai = PatternFill("solid", fgColor="F4AF85")
    fill_cw = PatternFill("solid", fgColor="FED966")

    line = 3
    eline=2
    for patientid,rpos_list in cw_dict.items():
        ai_rpos_list = ai_dict[patientid]
        ai_match=[]
        for rpos in rpos_list:
            sheet.cell(row = line ,column= 1).value=patientid
            sheet.cell(row=line, column=2).value = rpos
            for ai_pos in ai_rpos_list:
                if abs(ai_pos-rpos)<36:
                    sheet.cell(row=line, column=3).value = ai_pos
                    ai_match.append(ai_pos)
            line+=1

        ai_error = list(set(ai_rpos_list)-set(ai_match))
        ai_error.sort()

        if bool(ai_error):
            for epos in ai_error:
                sheet2.cell(row=eline, column=1).value = patientid
                sheet2.cell(row=eline, column=2).value = epos
                eline+=1


    wb.save(date + 'ai_carewell_4.2sQRS心搏验证结果3.xlsx')


def match_excel_result():
    # wb=openpyxl.Workbook()
    wb = openpyxl.load_workbook('统计模板.xlsx')
    sheet_ai = wb.get_sheet_by_name('张雪patient_match')
    sheet_cw = wb.get_sheet_by_name('田亮patient_match')
    # sheet_num = wb.get_sheet_by_name('个数-result')
    cw_match, ai_match, cw_error, ai_error, cw_match_patient_num, ai_match_patient_num = match_label_resucw()

    matchline_ai = 2
    matchline_cw = 2
    for pid, matchlist in cw_match.items():
        for matchpr in matchlist:
            sheet_cw.cell(row=matchline_cw, column=1).value = pid
            sheet_cw.cell(row=matchline_cw, column=2).value = matchpr[0]
            sheet_cw.cell(row=matchline_cw, column=3).value = matchpr[1]
            matchline_cw += 1

    for pid, matchlist in ai_match.items():
        for matchpr in matchlist:
            sheet_ai.cell(row=matchline_ai, column=1).value = pid
            sheet_ai.cell(row=matchline_ai, column=2).value = matchpr[0]
            sheet_ai.cell(row=matchline_ai, column=3).value = matchpr[1]
            matchline_ai += 1

    wb.save('张雪田亮4.2sQRS心搏验证结果20171211-66.xlsx')


def error_excel_result():
    # wb=openpyxl.Workbook()
    wb = openpyxl.load_workbook('统计模板.xlsx')
    sheet_ai = wb.get_sheet_by_name('张雪patient_error')
    sheet_cw = wb.get_sheet_by_name('田亮patient_error')
    # sheet_num = wb.get_sheet_by_name('个数-result')
    cw_match, ai_match, cw_error, ai_error, cw_match_patient_num, ai_match_patient_num = match_label_resucw()

    # error_line = 2
    # for pid,



    errorline_ai = 2
    errorline_cw = 2
    for pid, errorlist in cw_error.items():
        for errorpr in errorlist:
            sheet_cw.cell(row=errorline_cw, column=1).value = pid
            sheet_cw.cell(row=errorline_cw, column=2).value = errorpr[1]
            sheet_cw.cell(row=errorline_cw, column=3).value = errorpr[0]
            sheet_cw.cell(row=errorline_cw, column=4).value = errorpr[2]
            errorline_cw += 1

    for pid, errorlist in ai_error.items():
        for errorpr in errorlist:
            sheet_ai.cell(row=errorline_ai, column=1).value = pid
            sheet_ai.cell(row=errorline_ai, column=2).value = errorpr[1]
            sheet_ai.cell(row=errorline_ai, column=3).value = errorpr[0]
            sheet_ai.cell(row=errorline_ai, column=4).value = errorpr[2]
            errorline_ai += 1

    wb.save('张雪田亮4.2sQRS心搏验证结果20171211-67.xlsx')


if __name__ == '__main__':
    date = '20180102'
    get_result(date)
