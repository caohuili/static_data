# coding=utf-8
import re,os
import pymysql
import collections
import pandas as pd
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill,Alignment


def get_data():
    # conn = MySQLdb.connect(host='10.10.9.156',port = 3306,user = 'caohuili',password = 'chl##098',db = 'StaticECG')
    conn = pymysql.connect(host='10.10.9.156', port=3306, user='caohuili', password='chl##098', db='FDA_Carewell',
                           charset='utf8')
    cur = conn.cursor()
    sql_qiang = '''select PatientID,DIID from t_ecg_event WHERE PatientID IN (SELECT PatientID FROM `t_test` WHERE TestType=2 or TestType=3);'''
    sql_carewell = 'select PatientID,DoctorAdvice from t_test WHERE TestType=2 or TestType=3'
    sql_qiang_beizhu='select PatientID,CustomDiagnosis from t_patient'

    # sql_qiang = '''select PatientID,DIID from t_ecg_event WHERE PatientID in (select PatientID from t_patient WHERE DataSource=2)'''
    # sql_carewell = 'select PatientID,DoctorAdvice from t_test WHERE TestType=18 and PatientID in (select PatientID from t_patient WHERE DataSource=2)'
    # sql_qiang_beizhu='select PatientID,CustomDiagnosis from t_patient WHERE PatientID in (select PatientID from t_patient WHERE DataSource=2)'

    cur.execute(sql_qiang)
    cur.execute(sql_carewell)
    cur.execute(sql_qiang_beizhu)

    a=pd.read_sql(sql_carewell, conn)
    df_qiang = pd.read_sql(sql_qiang, conn).drop_duplicates().dropna()
    df_carewell = pd.read_sql(sql_carewell, conn).drop_duplicates().dropna()
    df_qiang_beizhu = pd.read_sql(sql_qiang_beizhu, conn).dropna()

    df_qiang.sort_values('PatientID')
    df_carewell.sort_values('PatientID')
    # print(len(df_qiang),len(df_carewell),len(a))

    # l1=set(a['PatientID'])
    # l2=set(df_carewell['PatientID'])
    # print(l1-l2)

    cur.close()
    conn.commit()
    conn.close()
    return df_qiang, df_carewell,df_qiang_beizhu

def get_excel_result():

    wb=openpyxl.load_workbook('诊断统计.xlsx')
    sheet=wb.get_sheet_by_name('result')

    df_qiang, df_carewell,df_qiang_beizhu = get_data()

    patientids = list(set(df_qiang['PatientID']))
    patientids.sort()

    # print(df_carewell[df_carewell['PatientID']==282006])
    line=2
    for patientid in patientids:
        #print(patientid)
        q_code_list = list(df_qiang[df_qiang['PatientID']==patientid]['DIID'])
        try:
            c_code = list(df_carewell[df_carewell['PatientID']==patientid]['DoctorAdvice'])[0]
        except:
            continue
        qiang_beizhu =list(df_qiang_beizhu[df_qiang_beizhu['PatientID'] == patientid]['CustomDiagnosis'])

        c_code_text = '|'+str(c_code)
        #print(q_code_list,c_code_text.replace('||','|').replace(': ',':'))
        c_code_text=c_code_text.replace('||', '|').replace(': ', ':').replace('    ','')
        c_code_text_list = c_code_text.split('|')
        #print(c_code_text_list)
        c_code_list=[]
        c_code_detail=[]
        for carewell_code in c_code_text_list:
            if carewell_code == ''or carewell_code=='None':
                continue
            #print(carewell_code.split(':')[0])
            c_code_list.append(int(carewell_code.split(':')[0]))
            try:
                c_code_detail.append(carewell_code.split(':')[2])
            except:
                c_code_detail.append('')

        #print(c_code_list)
        if 875 in c_code_list:
            c_code_list = c_code_list+[504,871,872]
        else:
            pass
        codes_list = list(set(q_code_list+c_code_list))
        codes_list.sort()
        col=2
        for code_id in codes_list:
            sheet.cell(row=line,column=1).value=patientid
            sheet.cell(row=line, column=2).value = code_id
            if code_id in q_code_list:
                sheet.cell(row=line, column=3).value = code_id

            if code_id in c_code_list:
                sheet.cell(row=line, column=4).value = code_id
                sheet.cell(row=line, column=5).value = c_code_detail[c_code_list.index(code_id)]
            line+=1
        # if bool(qiang_beizhu):
        #     sheet.cell(row=line,column=1).value=patientid
        #     sheet.cell(row=line, column=3).value = qiang_beizhu[0]
        #     line+=1
    sheet.freeze_panes = 'A2'
    wb.save('20180126_强大夫_carewell_诊断结果统计-1.xlsx')


def get_excel_patient_result(page,codeid_list,patient_excel):
    more_match_set = {504,871,872,875}

    wb=openpyxl.load_workbook('sen-pos-%d.xlsx'%(page-1))
    sheet=wb.get_sheet_by_name('patientid-%d'%page)

    df_qiang, df_carewell,df_qiang_beizhu = get_data()

    codeid_dict = collections.defaultdict(list)
    for codeid in codeid_list:
        codeid_dict[codeid]=[0, 0, 0]

    patientids = list(set(df_qiang['PatientID']))
    patientids.sort()
    total_patient = len(patientids)

    line=4

    for patientid in patientids:
        q_code_list = list(df_qiang[df_qiang['PatientID']==patientid]['DIID'])
        try:
            c_code = list(df_carewell[df_carewell['PatientID']==patientid]['DoctorAdvice'])[0]
        except:
            total_patient=total_patient-1
            continue
        qiang_beizhu =list(df_qiang_beizhu[df_qiang_beizhu['PatientID'] == patientid]['CustomDiagnosis'])

        c_code_text = '|'+str(c_code)
        c_code_text=c_code_text.replace('||', '|').replace(': ', ':').replace('    ','')

        c_code_list=re.findall(r'\|(\d\d\d):',c_code_text)
        c_code_set=set(map(int,c_code_list))
        if 875 in c_code_set:
            c_code_set = c_code_set|more_match_set
        else:
            pass

        match = set(q_code_list)&c_code_set
        q_less = c_code_set-set(q_code_list)
        q_more = set(q_code_list)-c_code_set

        col=1
        for codeid in codeid_list:
            if codeid in match:
                codeid_dict[codeid][0]+=1
                sheet.cell(row=codeid_dict[codeid][0]+3, column=col).value = patientid
            if codeid in q_more:
                codeid_dict[codeid][1] += 1
                sheet.cell(row=codeid_dict[codeid][1]+3, column=col+1).value = patientid
            if codeid in q_less:
                codeid_dict[codeid][2] += 1
                sheet.cell(row=codeid_dict[codeid][2]+3, column=col+2).value = patientid
            col+=3

    print(total_patient)
    col_num = 1

    label = ['TP','FN','FP']
    for codeid in codeid_list:
        code_num_list = codeid_dict[codeid]
        sheet.merge_cells(start_row = 1,end_row=1,start_column=col_num,end_column=col_num+2)
        sheet.cell(row=1,column=col_num).value=codeid
        for i in range(3):
            sheet.cell(row=2, column=i+col_num).value = label[i]
            sheet.cell(row=3, column=col_num+i).value = code_num_list[i]
        col_num+=3
    sheet.freeze_panes='A2'
    wb.save(patient_excel)
    print(total_patient)
    return total_patient,codeid_dict
    #wb.save('20180125_强大夫_carewell_诊断结果个数统计3.xlsx')

def get_excel_senpos_result(page,codeid_list,patient_excel,sen_pos_excel):

    total_patient, codeid_dict = get_excel_patient_result(page,codeid_list,patient_excel)

    wb=openpyxl.load_workbook(patient_excel)
    sheet=wb.get_sheet_by_name('sen-pos-%d'%page)

    line=2
    for codeid in codeid_list:
        sheet.cell(row=line, column=2).value = codeid
        code_num_list = codeid_dict[codeid]

        tp=code_num_list[0]
        fn = code_num_list[1]
        fp = code_num_list[2]
        tn = total_patient-tp-fn-fp

        # sens=tp/(tp+fn)
        # spec=tn/(tn+fp)
        # ppv = tp/(tp+fp)
        # npv=tn/(tn+fn)
        # prev = tp/total_patient

        # sens=zero_division(tp,(tp+fn))
        # spec=zero_division(tn,(tn+fp))
        # ppv = zero_division(tp,(tp+fp))
        # npv=zero_division(tn,(tn+fn))
        # prev_div = '%d/%d'%(tp+fn,total_patient)
        # prev = zero_division(tp+fn,total_patient)

        all_num_list = [tp, tn, fp, fn, total_patient]
        col = 3

        for num in all_num_list:
            sheet.cell(row=line, column=col).value = num
            col+=1
            if col==8:

                sheet.cell(row=line, column=8).value = zero_division('=round(C%d*100/(C%d+F%d),1)' % (line, line, line), tp, (tp + fn))
                sheet.cell(row=line, column=9).value = zero_division('=round(D%d*100/(D%d+E%d),1)' % (line, line, line), tn, (tn + fp))
                sheet.cell(row=line, column=10).value = zero_division('=round(C%d*100/(C%d+E%d),1)' % (line, line, line), tp, (tp + fp))
                sheet.cell(row=line, column=11).value = zero_division('=round(D%d*100/(D%d+F%d),1)' % (line, line, line), tn, (tn + fn))
                sheet.cell(row=line, column=12).value = '=SUM(C%d,F%d)&"/"&G%d' % (line, line, line)
                sheet.cell(row=line, column=12).alignment = Alignment(horizontal='right',vertical='center')
                sheet.cell(row=line, column=13).value = zero_division('=round(SUM(C%d,F%d)*100/G%d,1)' % (line, line, line), tp + fn, total_patient)
            # elif col == 4:
            #     sheet.cell(row=line, column=4).value = '=G%d-SUM(C%d,E%d,F%d)'%(line,line,line,line)
        sheet.cell(row=line, column=4).value = '=G%d-SUM(C%d,E%d,F%d)' % (line, line, line, line)
        line+=1

        #all_num_list = [tp,tn,fp,fn,total_patient,sens,spec,ppv,npv,prev_div,prev]
        # all_num_list = [tp, tn, fp, fn, total_patient]
        # col = 2
        # for num in all_num_list:
        #     sheet.cell(row=line, column=col).value = num
        #     col+=1
        #
        # line+=1
    sheet.freeze_panes = 'A2'
    wb.save(sen_pos_excel)
    if os.path.exists(patient_excel):
        os.remove(patient_excel)


def zero_division(s,x,y):
    try:
        z=x/y
    except:
        s='-'
    return s
if __name__ == '__main__':
    patient_excel ='20180222_强大夫_carewell_诊断结果个数统计.xlsx'
    # sen_pos_excel ='sen-pos-%d.xlsx'%num
    get_data()
    # codeid_list1 = [504,505,511,501,502,402,510]
    # codeid_list2=[110,811,871,821,812,822,872,804,823,803,868,421]
    # codeid_list3 =[842,841,410,414,845,846,869,412,415,848,413]
    # codeid_list4=[201,202,203,204,205,206,701]
    # #codeid_list4=[701]
    # #codeid_list = [ codeid_list4]
    # codeid_list=[codeid_list1,codeid_list2,codeid_list3,codeid_list4]
    # #get_excel_result()
    # #get_excel_patient_result(codeid_list)
    # for num in range(1,5):
    #     sen_pos_excel = 'sen-pos-%d.xlsx' % num
    #     get_excel_senpos_result(num,codeid_list[num-1],patient_excel,sen_pos_excel)
