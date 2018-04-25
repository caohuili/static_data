import openpyxl
from openpyxl.styles import PatternFill

def get_label():
    wb = openpyxl.load_workbook('E:\\xindian\\静态心电数据库统计\\静态心搏统计\\静态心搏统计.xlsx')
    sheet_zhx = wb.get_sheet_by_name('1000')
    sheet_tl = wb.get_sheet_by_name('500')
    sheet_q = wb.get_sheet_by_name('q')
    zhx_lines = sheet_zhx.max_row
    tl_lines = sheet_tl.max_row
    q_lines = sheet_q.max_row
    zhx = {}
    tl={}
    q = {}

    for i in range(2, zhx_lines + 1):
        patientid = sheet_zhx.cell(row=i, column=1).value
        rpos = sheet_zhx.cell(row=i, column=2).value
        rlabel = sheet_zhx.cell(row=i, column=3).value
        if patientid in zhx:

            pass
        else:
            zhx[patientid] = {}
        zhx[patientid][rpos] = rlabel
        
    for i in range(2, tl_lines + 1):
        patientid = sheet_tl.cell(row=i, column=1).value
        rpos = sheet_tl.cell(row=i, column=2).value
        rlabel = sheet_tl.cell(row=i, column=3).value
        if patientid in tl:

            pass
        else:
            tl[patientid] = {}
        tl[patientid][rpos] = rlabel

    for i in range(2, q_lines + 1):
        patientid = sheet_q.cell(row=i, column=1).value
        rpos = sheet_q.cell(row=i, column=2).value
        rlabel = sheet_q.cell(row=i, column=3).value
        if patientid in q:
            pass
        else:
            q[patientid] = {}
        q[patientid][rpos] = rlabel

    return zhx,tl, q


def match_label_resutl(date):
    zhx_dict, tl_dict,q_dict = get_label()
    zhx_match = {}
    tl_match = {}
    zhx_error = {}
    tl_error={}
    # all_label=['N', 'Af','AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'Af_LVH', 'N_CLB', 'Af_CRB', 'N_RB', 'V', 'Af_CLB', 'SE',
    # 'N_LAE', 'N_PS', 'JE', 'A', 'VE', 'AT']
    all_label = ['N', 'Af', 'AF', 'N_LVH', 'N_B1', 'N_CRB', 'S', 'N_CLB', 'Af_CRB', 'V', 'SE',
                 'N_PS', 'JE', 'A', 'VE', 'AT', 'VT']
    total = {}
    match_label = {}

    match_patient_num =0
    zhx_match_patient_num = 0
    tl_match_patient_num = 0

    for lab in all_label:
        total[lab] = 0
        match_label[lab] = 0
    for patientid, label_dict in q_dict.items():

        if patientid in zhx_dict and patientid in tl_dict:
            # tl_match[patientid] = []
            # zhx_match[patientid] = []
            # tl_error[patientid] = []
            # zhx_error[patientid] = []


            zhx_label_dict = zhx_dict[patientid]
            tl_label_dict = tl_dict[patientid]
            zhx_match_list=[]
            zhx_error_list=[]
            zhx=[]
            tl=[]
            tl_match_list=[]
            tl_error_list=[]
            
            zhxpos_list = list(zhx_label_dict.keys())
            zhxpos_list.sort()
            #print(zhxpos_list)
            #for zhxpos, zhxlabel in zhx_label_dict.items():
            for zhxpos in zhxpos_list:
                zhxlabel=zhx_label_dict[zhxpos]
                for pos, label in label_dict.items():
                    if abs(zhxpos - pos) < 101:
                        #print(zhxlabel,label)
                        if label == zhxlabel:
                            zhx.append([zhxpos, zhxlabel])
                            zhx_error_list.append([label,zhxpos, zhxlabel])
                        else:
                            zhx_error_list.append([label,zhxpos, zhxlabel])
                    else:
                        pass
            if len(zhx)==len(zhx_label_dict):
                zhx_match_patient_num+=1
                zhx_match[patientid]=zhx
            else:
                zhx_error[patientid] = zhx_error_list

            tlpos_list = list(tl_label_dict.keys())
            tlpos_list.sort()
            # for tlpos, tllabel in tl_label_dict.items():
            for tlpos in tlpos_list:
                tllabel = tl_label_dict[tlpos]
                for pos, label in label_dict.items():
                    if abs(tlpos - pos) < 101:
                        if label == tllabel:
                            tl.append([tlpos, tllabel])
                            tl_error_list.append([label,tlpos, tllabel])
                        else:
                            tl_error_list.append([label,tlpos, tllabel])
                    else:
                        pass
            if len(tl)==len(tl_label_dict):
                tl_match_patient_num+=1
                tl_match[patientid]=tl
            else:
                tl_error[patientid] = tl_error_list
    # print(set(tl_error))
    # print(len(tl_error))
    print(tl_match_patient_num, zhx_match_patient_num)

    #return tl_match, zhx_match, tl_error, zhx_error, tl_match_patient_num, zhx_match_patient_num
    #########################################################################################
    zhx_only_match_list = list(set(zhx_match)&set(tl_error))
    zhx_only_match_list.sort()

    #print(zhx_only_match_list)
    tl_only_match_list = list(set(tl_match) & set(zhx_error))
    tl_only_match_list.sort()

    zhx_tl_error_list = list(set(zhx_error)|set(tl_error))
    zhx_tl_error_list.sort()

    #return tl_match, zhx_match, tl_error, zhx_error,zhx_only_match_list,tl_only_match_list,zhx_tl_error_list

    get_three_people_result(tl_match, zhx_match, tl_error, zhx_error, zhx_only_match_list, tl_only_match_list, zhx_tl_error_list,date)


def get_three_people_result(tl_match, zhx_match, tl_error, zhx_error, zhx_only_match_list, tl_only_match_list, zhx_tl_error_list,date):
    # wb=openpyxl.Workbook()
    #tl_match, zhx_match, tl_error, zhx_error, zhx_only_match_list, tl_only_match_list, zhx_tl_error_list= match_label_resutl()
    wb = openpyxl.load_workbook('统计模板2.xlsx')
    sheet_zhx = wb.get_sheet_by_name('张雪match田亮notmatch')
    sheet_tl = wb.get_sheet_by_name('田亮match张雪notmatch')
    sheet_notmatch = wb.get_sheet_by_name('不匹配patient对比')

    fill_zhx = PatternFill("solid", fgColor="F4AF85")
    fill_tl = PatternFill("solid", fgColor="FED966")

    matchline_zhx = 3
    matchline_tl = 3
    notmatchline=2

    for pid in zhx_only_match_list:
        zhx_only_matchlist=zhx_match[pid]
        for i in range(len(zhx_only_matchlist)):
            matchpr = zhx_only_matchlist[i]
            sheet_zhx.cell(row=matchline_zhx, column=1).value = pid
            sheet_zhx.cell(row=matchline_zhx, column=2).value = matchpr[0]
            sheet_zhx.cell(row=matchline_zhx, column=3).value = matchpr[1]
            sheet_zhx.cell(row=matchline_zhx, column=4).value = matchpr[1]
            # print(tl_error[pid])
            # print(i)
            # print(tl_error[pid][i])
            try:
                tl_rlabel=tl_error[pid][i][2]
            except:
                tl_rlabel = ''
            if tl_rlabel!=matchpr[1]:
                sheet_zhx.cell(row=matchline_zhx, column=5).fill=fill_tl
            else:
                pass
            sheet_zhx.cell(row=matchline_zhx, column=5).value = tl_rlabel
            matchline_zhx += 1

    for pid in tl_only_match_list:
        tl_only_matchlist=tl_match[pid]

        for i in range(len(tl_only_matchlist)):
            matchpr=tl_only_matchlist[i]
            sheet_tl.cell(row=matchline_tl, column=1).value = pid
            sheet_tl.cell(row=matchline_tl, column=2).value = matchpr[0]
            sheet_tl.cell(row=matchline_tl, column=3).value = matchpr[1]
            sheet_tl.cell(row=matchline_tl, column=4).value = matchpr[1]
            zhx_rlabel = zhx_error[pid][i][2]
            if zhx_rlabel!=matchpr[1]:
                sheet_tl.cell(row=matchline_tl, column=5).fill=fill_zhx
            else:
                pass
            sheet_tl.cell(row=matchline_tl, column=5).value = zhx_rlabel
            matchline_tl += 1

    for pid in zhx_tl_error_list:
        if pid in tl_error:
            tl_error_label = tl_error[pid]
            for i in range(len(tl_error_label)):
                label=tl_error_label[i]
                sheet_notmatch.cell(row=notmatchline, column=1).value = pid
                sheet_notmatch.cell(row=notmatchline, column=2).value = label[1]
                sheet_notmatch.cell(row=notmatchline, column=3).value = label[0]
                sheet_notmatch.cell(row=notmatchline, column=4).value = label[2]
                if pid in zhx_match:
                    sheet_notmatch.cell(row=notmatchline, column=5).value = zhx_match[pid][i][1]
                else:
                    sheet_notmatch.cell(row=notmatchline, column=5).value = zhx_error[pid][i][2]

                if sheet_notmatch.cell(row=notmatchline, column=5).value!=label[0]:
                    sheet_notmatch.cell(row=notmatchline, column=5).fill=fill_zhx
                if label[2]!=label[0]:
                    sheet_notmatch.cell(row=notmatchline, column=4).fill = fill_tl


                notmatchline += 1
        else:
            zhx_error_label = zhx_error[pid]
            for i in range(len(zhx_error_label)):
                label = zhx_error_label[i]
                sheet_notmatch.cell(row=notmatchline, column=1).value = pid
                sheet_notmatch.cell(row=notmatchline, column=2).value = label[1]
                sheet_notmatch.cell(row=notmatchline, column=3).value = label[0]
                sheet_notmatch.cell(row=notmatchline, column=5).value = label[2]
                try:

                    sheet_notmatch.cell(row=notmatchline, column=4).value = tl_match[pid][i][1]
                except:
                    sheet_notmatch.cell(row=notmatchline, column=4).value = ''

                if label[2]!=label[0]:
                    sheet_notmatch.cell(row=notmatchline, column=5).fill = fill_zhx

                notmatchline += 1


    wb.save(date+'ai_carewell_4.2sQRS心搏验证结果.xlsx')


def match_excel_result():
    # wb=openpyxl.Workbook()
    wb = openpyxl.load_workbook('统计模板.xlsx')
    sheet_zhx = wb.get_sheet_by_name('张雪patient_match')
    sheet_tl = wb.get_sheet_by_name('田亮patient_match')
    #sheet_num = wb.get_sheet_by_name('个数-result')
    tl_match, zhx_match, tl_error, zhx_error, tl_match_patient_num, zhx_match_patient_num = match_label_resutl()

    matchline_zhx = 2
    matchline_tl = 2
    for pid, matchlist in tl_match.items():
        for matchpr in matchlist:
            sheet_tl.cell(row=matchline_tl, column=1).value = pid
            sheet_tl.cell(row=matchline_tl, column=2).value = matchpr[0]
            sheet_tl.cell(row=matchline_tl, column=3).value = matchpr[1]
            matchline_tl += 1
            
    for pid, matchlist in zhx_match.items():
        for matchpr in matchlist:
            sheet_zhx.cell(row=matchline_zhx, column=1).value = pid
            sheet_zhx.cell(row=matchline_zhx, column=2).value = matchpr[0]
            sheet_zhx.cell(row=matchline_zhx, column=3).value = matchpr[1]
            matchline_zhx += 1

    wb.save('张雪田亮4.2sQRS心搏验证结果20180327.xlsx')


def error_excel_result(date):
    # wb=openpyxl.Workbook()
    wb = openpyxl.load_workbook('统计模板.xlsx')
    sheet_zhx = wb.get_sheet_by_name('张雪patient_error')
    sheet_tl = wb.get_sheet_by_name('田亮patient_error')
    # sheet_num = wb.get_sheet_by_name('个数-result')
    tl_match, zhx_match, tl_error, zhx_error, zhx_only_match_list, tl_only_match_list, zhx_tl_error_list
    tl_match, zhx_match, tl_error, zhx_error, tl_match_patient_num, zhx_match_patient_num = match_label_resutl(date)

    # error_line = 2
    # for pid,



    errorline_zhx = 2
    errorline_tl = 2
    for pid, errorlist in tl_error.items():
        for errorpr in errorlist:
            sheet_tl.cell(row=errorline_tl, column=1).value = pid
            sheet_tl.cell(row=errorline_tl, column=2).value = errorpr[1]
            sheet_tl.cell(row=errorline_tl, column=3).value = errorpr[0]
            sheet_tl.cell(row=errorline_tl, column=4).value = errorpr[2]
            errorline_tl += 1

    for pid, errorlist in zhx_error.items():
        for errorpr in errorlist:
            sheet_zhx.cell(row=errorline_zhx, column=1).value = pid
            sheet_zhx.cell(row=errorline_zhx, column=2).value = errorpr[1]
            sheet_zhx.cell(row=errorline_zhx, column=3).value = errorpr[0]
            sheet_zhx.cell(row=errorline_zhx, column=4).value = errorpr[2]
            errorline_zhx += 1

    wb.save('张雪田亮4.2sQRS心搏验证结果20180327-2.xlsx')

if __name__ == '__main__':
    date='20180327'
    # get_label()
    # label_resutl()
    #match_excel_result()
    #match_label_resutl(date)
    error_excel_result(date)
    #match_excel_result
